import pandas as pd
import numpy as np
import random
import csv
import openpyxl
from openpyxl import load_workbook


Rhonda_Wins = ["Get a bigger closet", "Travel the world", "Take a nap", "Gamble in Wendover", "Buy orthopedic shoes", "Drink prune juice", "Relax", "Catch a matinee", "Read a book", "Eat more fiber", "Count your money", "Watch game shows", "Take a long bath", "Get a massage", "Go for a walk", "Go to the beach", "See a show in Vegas", "Eat Cake!", "Take advange of a senior discount", "Do nothing", "take a long lunch"]
Bad = ["Play Brain Games", "Get financially savvy", "establish a routine", "Take on a new hobby", "Work part time", "Declutter your home", "Find your purpose", "Live in the moment", "Discover local attractions", "Become a tour guide", "Focus on your Personal Style", "Find a solo activity", "Become a birder", "Learn to dance", "Get REALLY into WWII", "Sew a quilt", "Reminisce", "Build a bird house", "Sleep in"]

def build_bingo_card(isWinningCard, total_bingos):
    #build the card
    card = np.zeros((5,5))
    
    #set the free space
    card[2][2] = True

    if isWinningCard:
        #put the bingos on the card
        add_bingos(card, total_bingos)
    else:
        build_crap_card(card, total_bingos)

    print(card)
    print("")

    return card



def add_bingos(card, total_bingos):
    for i in range(total_bingos):
        make_bingo_combo(card)
    return

def make_bingo_combo(card):
    length = len(card)
    choice = random.choice(["row", "col", "diag_l", "diag_r"])
    if choice == "row":
        chosen_row = random.randint(0, 4)
        for j in range(len(card[chosen_row])):
            card[chosen_row][j] = True
    if choice == "col":
        chosen_col = random.randint(0, 4)
        for j in range(len(card)):
            card[j][chosen_col] = True
    if choice == "diag_l":
        for i in range(length):
            card[i][i] = True
    if choice == "diag_r":
        for i in range(length):
            card[i][length-1-i] = True
    return


def build_crap_card(card, total_bingos):
    true_limit = calculate_bingo_area(total_bingos)
    for i in range(true_limit-2):
        row = random.randint(0, 4)
        col = random.randint(0, 4)
        card[row][col] = True
    verify_bad_card(card)    
    # return card

def verify_bad_card(card):
    #for the rows
    for i in range(len(card)):
        if sum(card[i]) == 5:
            card[i][random.randint(0, 4)] = False
    
    #columns
    for i in range(len(card[0])):    
        sum_col = 0
        for j in range(len(card)):
            sum_col += card[j][i]
        if sum_col == 5:
            card[i][4] = False
            sum_col = 0

    #for the diags
    card[0][0] = False
    card[3][1]= False

def calculate_bingo_area(total_bingos):
    return (5*total_bingos) - (total_bingos*1)


def write_card_to_csv(card, coords, input_file, write_solution=False):
    x = coords[0]
    y = coords[1]

    my_file = open(input_file, "rb")
    solution = open("solution.txt", "a")
    my_workbook = openpyxl.load_workbook(my_file)
    worksheet = my_workbook.active
    rhonda_wins = Rhonda_Wins.copy()
    bad = Bad.copy()

    for i in range(len(card)):
        for j in range(len(card)):
            row = i+x
            col = j+y
            cell = worksheet.cell(row=row, column=col)
            if i==2 and j==2:
                cell.value = "FREE SPACE"
                continue
            if card[i][j] == True:
                cell.value = pick_and_remove_option(cell, rhonda_wins)
                if write_solution:
                    solution.write(cell.value + '\n')
            else:
                cell.value = pick_and_remove_option(cell, bad)

    solution.close()
    my_workbook.save(input_file)
    return

def pick_and_remove_option(cell, options_list):
    value = random.sample(options_list, 1)
    options_list.remove(value[0])
    return  value[0]

def build_bingo_sheet(has_Rhonda, my_file):
    starting_coords = [(4,1), (4,7), (14,1), (14,7), (24,1), (24,7)]
    total_bingos = 4
    if has_Rhonda:
        winning_card = build_bingo_card(True, 5)
        write_card_to_csv(winning_card, starting_coords[0], my_file, True)
        for i in range(1,6):
            card = build_bingo_card(False, total_bingos)
            write_card_to_csv(card, starting_coords[i], my_file)
    else:
        for i in range(len(starting_coords)):
            card = build_bingo_card(False, total_bingos)



build_bingo_sheet(True, "Rhondas Retirement 1st one is WINNING SHEET.xlsx")
for i in range(6):
    file_name = "Rhondas Retirement BAD CARDS_{0}.xlsx".format(i)
    build_bingo_sheet(False,file_name)

# build_bingo_card(True, 5)
